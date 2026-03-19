# app.py
# PharmaROI Intelligence — V3 (Multi-Model Comparison)
# Run: streamlit run "PharmaROI Model/app_v3.py"

from __future__ import annotations

import copy
import io
from dataclasses import dataclass
from typing import List

import streamlit as st

try:
    import pandas as pd
except Exception:
    pd = None

import plotly.express as px
import plotly.graph_objects as go

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import github_storage as gs

# -----------------------------
# Color palette
# -----------------------------
COLORS = {
    "primary": "#0F6CBD",
    "revenue": "#0F6CBD",
    "costs": "#9CA3AF",
    "profit": "#10B981",
    "warning": "#F59E0B",
    "danger": "#EF4444",
    "muted": "#6B7280",
}

TAB_PALETTE = [
    "#0F6CBD", "#10B981", "#F59E0B", "#EF4444",
    "#8B5CF6", "#EC4899", "#06B6D4", "#84CC16",
]

# -----------------------------
# Funnel definitions
# -----------------------------
STAGE_NAMES: List[str] = [
    "Total Addressable Market for MASH",
    "F2 and F3",
    "MASH patients diagnosed",
    "Madrigal access to MASH patients",
    "Frequent users of online and social media resources",
    "Activation within 90 days mo onto Dario Connect for MASH",
    "Schedule telemedicine appointment",
    "Keep telemedicine appointment",
    "Obtain prescription for biopsy",
    "Get biopsy lab test",
    "Get positive lab results",
    "Complete post lab result consultation",
    "Get prescription for Rezdiffra",
]

SPONSOR_DEFAULTS = {
    "base_population": 10_000_000,
    "ratios": [1.00, 0.35, 0.16, 0.22, 0.75, 0.40, 0.15, 0.80, 1.00, 0.75, 0.90, 0.50, 0.90],
    "cac": [0.0, 0.0, 0.0, 0.0, 0.0, 10.0, 67.0, 83.0, 83.0, 111.0, 123.0, 247.0, 274.0],
    "arpp": 47_400.0,
    "treatment_years": 1.0,
    "discount": 0.68,
    "stage_active": [True] * len(STAGE_NAMES),
    "stage_names": STAGE_NAMES[:],
    "platform_costs": {
        "dario_connect_config": 500_000.0,
        "dario_care_config": 500_000.0,
        "sub_dario_connect": 1_000_000.0,
        "sub_dario_care": 250_000.0,
        "maintenance_support": 250_000.0,
    },
}

ZERO_SAMPLE = {
    "base_population": 0,
    "ratios": [0.0] * len(STAGE_NAMES),
    "cac": [0.0] * len(STAGE_NAMES),
    "arpp": 0.0,
    "treatment_years": 1.0,
    "discount": 0.0,
    "stage_active": [True] * len(STAGE_NAMES),
    "stage_names": ["Insert Stage Name"] * len(STAGE_NAMES),
    "platform_costs": {
        "dario_connect_config": 0.0,
        "dario_care_config": 0.0,
        "sub_dario_connect": 0.0,
        "sub_dario_care": 0.0,
        "maintenance_support": 0.0,
    },
}

# -----------------------------
# Formatting helpers
# -----------------------------
def clamp(x, lo, hi):
    return max(lo, min(hi, float(x)))

def money(x):
    return f"${x:,.0f}"

def number(x):
    return f"{x:,.0f}"

def pct(x):
    return f"{x*100:,.1f}%"

def roix(x):
    return f"{x:,.2f}x"

# -----------------------------
# Core computations
# -----------------------------
@dataclass(frozen=True)
class StageInput:
    name: str
    active: bool
    ratio: float
    cac: float

@dataclass(frozen=True)
class StageResult:
    name: str
    active: bool
    ratio_used: float
    patients: float
    cac_per_patient: float
    stage_cac: float
    cumulative_cac: float


def compute_funnel(stages: List[StageInput], base_population: float) -> List[StageResult]:
    results = []
    prev_patients = max(0.0, float(base_population))
    total_cac_pool = 0.0

    for idx, s in enumerate(stages):
        if idx == 0:
            patients = prev_patients
            ratio_used = 1.0
        else:
            ratio_used = 1.0 if not s.active else clamp(s.ratio, 0.0, 1.0)
            patients = prev_patients * ratio_used

        if idx < 5:
            cac_pp = 0.0
            stage_cac = 0.0
            cumulative = 0.0
        elif idx == 5:
            cac_pp = 0.0 if not s.active else max(0.0, float(s.cac))
            stage_cac = patients * cac_pp
            total_cac_pool = stage_cac
            cumulative = total_cac_pool
        else:
            cumulative = total_cac_pool
            cac_pp = (total_cac_pool / patients) if patients > 0 else 0.0
            stage_cac = cac_pp * patients

        results.append(StageResult(
            name=s.name, active=s.active, ratio_used=ratio_used,
            patients=patients, cac_per_patient=cac_pp,
            stage_cac=stage_cac, cumulative_cac=cumulative,
        ))
        prev_patients = patients

    return results


def compute_financials(treated_patients, arpp, treatment_years, discount, funnel_cac_total, platform_costs=0.0):
    treated = max(0.0, float(treated_patients))
    arpp = max(0.0, float(arpp))
    years = max(0.0, float(treatment_years))
    disc = clamp(discount, 0.0, 1.0)
    funnel_cac = max(0.0, float(funnel_cac_total))
    platform = max(0.0, float(platform_costs))

    gross = treated * arpp * years
    net = gross * (1.0 - disc)
    net_profit = net - funnel_cac - platform
    roi = (net / (funnel_cac + platform)) if (funnel_cac + platform) > 0 else float("nan")

    return {
        "treated_patients": treated,
        "gross_revenue": gross,
        "net_revenue": net,
        "discount": disc,
        "funnel_cac_total": funnel_cac,
        "platform_costs_total": platform,
        "net_profit": net_profit,
        "roi_net": roi,
    }


def run_model(state: dict):
    stage_names = state.get("stage_names", STAGE_NAMES)
    stages = []
    for idx, name in enumerate(stage_names):
        stages.append(StageInput(
            name=name,
            active=bool(state["stage_active"][idx]),
            ratio=float(state["ratios"][idx]) if idx > 0 else 1.0,
            cac=float(state["cac"][idx]),
        ))
    base_pop = float(state["base_population"])
    funnel_results = compute_funnel(stages, base_pop)
    platform_costs = sum(state.get("platform_costs", {}).values())
    fin = compute_financials(
        treated_patients=funnel_results[-1].patients,
        arpp=float(state["arpp"]),
        treatment_years=float(state["treatment_years"]),
        discount=float(state["discount"]),
        funnel_cac_total=funnel_results[-1].cumulative_cac,
        platform_costs=platform_costs,
    )
    return funnel_results, fin

# -----------------------------
# Monthly ROI helper
# -----------------------------
def build_monthly_roi_df(fin: dict, state: dict,
                         eff_0_3: float = 1.0,
                         eff_3_6: float = 1.0,
                         eff_6_plus: float = 1.0):
    months = max(1, int(round(float(state["treatment_years"]) * 12)))
    total_cost = float(fin["funnel_cac_total"] + fin["platform_costs_total"])
    monthly_net_revenue = float(fin["net_revenue"]) / months if months > 0 else 0.0

    rows = []
    cumulative_revenue = 0.0
    cumulative_cost = 0.0
    cumulative_profit = 0.0
    cumulative_phased_profit = 0.0
    phased_total_revenue = 0.0

    for month in range(1, months + 1):
        revenue = monthly_net_revenue
        cost = total_cost if month == 1 else 0.0

        if month <= 3:
            efficiency = eff_0_3
        elif month <= 6:
            efficiency = eff_3_6
        else:
            efficiency = eff_6_plus

        phased_revenue = revenue * efficiency
        phased_total_revenue += phased_revenue
        cumulative_phased_profit += phased_revenue - cost

        cumulative_revenue += revenue
        cumulative_cost += cost
        cumulative_profit = cumulative_revenue - cumulative_cost

        rows.append({
            "Month": month,
            "Monthly Net Revenue": revenue,
            "Monthly Cost": cost,
            "Cumulative Net Revenue": cumulative_revenue,
            "Cumulative Cost": cumulative_cost,
            "Cumulative Profit": cumulative_profit,
            "Cumulative Phased Profit": cumulative_phased_profit,
        })

    if pd is not None:
        df = pd.DataFrame(rows)
        payback_month = None
        positive = df[df["Cumulative Profit"] >= 0]
        if not positive.empty:
            payback_month = int(positive.iloc[0]["Month"])
        phased_roi = (phased_total_revenue / total_cost) if total_cost > 0 else float("nan")
        return df, payback_month, phased_total_revenue, phased_roi

    return rows, None, 0.0, float("nan")

# -----------------------------
# Sensitivity helper
# -----------------------------
def build_roi_sensitivity_df(state: dict, shock: float = 0.10):
    if pd is None:
        return None

    _, base_fin = run_model(copy.deepcopy(state))
    base_roi = base_fin["roi_net"]

    variables = [
        ("Base Population", ("scalar", "base_population")),
        ("ARPP", ("scalar", "arpp")),
        ("Discount", ("scalar", "discount")),
        ("Stage 2 Ratio", ("ratio", 1)),
        ("Stage 3 Ratio", ("ratio", 2)),
        ("Stage 4 Ratio", ("ratio", 3)),
        ("Stage 5 Ratio", ("ratio", 4)),
        ("Stage 6 Ratio", ("ratio", 5)),
        ("Final Stage Ratio", ("ratio", len(STAGE_NAMES) - 1)),
        ("Stage 6 CAC", ("cac", 5)),
        ("Platform Costs", ("platform_total", None)),
    ]

    rows = []

    for label, spec in variables:
        kind, key = spec
        low_state = copy.deepcopy(state)
        high_state = copy.deepcopy(state)

        if kind == "scalar":
            if key == "discount":
                low_state[key] = clamp(float(low_state[key]) * (1 - shock), 0.0, 1.0)
                high_state[key] = clamp(float(high_state[key]) * (1 + shock), 0.0, 1.0)
            else:
                low_state[key] = max(0.0, float(low_state[key]) * (1 - shock))
                high_state[key] = max(0.0, float(high_state[key]) * (1 + shock))

        elif kind == "ratio":
            idx = key
            low_state["ratios"][idx] = clamp(float(low_state["ratios"][idx]) * (1 - shock), 0.0, 1.0)
            high_state["ratios"][idx] = clamp(float(high_state["ratios"][idx]) * (1 + shock), 0.0, 1.0)

        elif kind == "cac":
            idx = key
            low_state["cac"][idx] = max(0.0, float(low_state["cac"][idx]) * (1 - shock))
            high_state["cac"][idx] = max(0.0, float(high_state["cac"][idx]) * (1 + shock))

        elif kind == "platform_total":
            for cost_key in low_state["platform_costs"]:
                low_state["platform_costs"][cost_key] = max(0.0, float(low_state["platform_costs"][cost_key]) * (1 - shock))
                high_state["platform_costs"][cost_key] = max(0.0, float(high_state["platform_costs"][cost_key]) * (1 + shock))

        _, low_fin = run_model(low_state)
        _, high_fin = run_model(high_state)

        low_delta = low_fin["roi_net"] - base_roi if (low_fin["roi_net"] == low_fin["roi_net"] and base_roi == base_roi) else 0.0
        high_delta = high_fin["roi_net"] - base_roi if (high_fin["roi_net"] == high_fin["roi_net"] and base_roi == base_roi) else 0.0

        rows.append({
            "Variable": label,
            "Low Delta": low_delta,
            "High Delta": high_delta,
            "Abs Impact": max(abs(low_delta), abs(high_delta)),
        })

    return pd.DataFrame(rows).sort_values("Abs Impact", ascending=True)

# -----------------------------
# Plotly chart helpers
# -----------------------------
def plotly_waterfall(fin):
    gross = fin["gross_revenue"]
    discount_amount = fin["gross_revenue"] - fin["net_revenue"]
    net_revenue = fin["net_revenue"]
    funnel_cac = fin["funnel_cac_total"]
    platform_costs = fin["platform_costs_total"]
    net_profit = fin["net_profit"]

    fig = go.Figure(go.Waterfall(
        name="Financial Bridge",
        orientation="v",
        measure=["relative", "relative", "total", "relative", "relative", "total"],
        x=["Gross Revenue", "Discount", "Net Revenue", "Funnel CAC", "Platform Costs", "Net Profit"],
        text=[money(gross), f"-{money(discount_amount)}", money(net_revenue), f"-{money(funnel_cac)}", f"-{money(platform_costs)}", money(net_profit)],
        textposition="outside",
        y=[gross, -discount_amount, 0, -funnel_cac, -platform_costs, 0],
        connector={"line": {"color": COLORS["muted"]}},
        increasing={"marker": {"color": COLORS["revenue"]}},
        decreasing={"marker": {"color": COLORS["danger"]}},
        totals={"marker": {"color": COLORS["profit"]}},
    ))
    fig.update_layout(height=420, margin=dict(l=10, r=10, t=40, b=10), showlegend=False, yaxis_title="USD")
    return fig

def plotly_sensitivity_tornado(sdf, shock: float = 0.10):
    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=sdf["Variable"], x=sdf["Low Delta"], orientation="h",
        name=f"-{int(shock*100)}%", marker_color=COLORS["danger"],
        hovertemplate="%{y}<br>ROI change: %{x:.2f}x<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        y=sdf["Variable"], x=sdf["High Delta"], orientation="h",
        name=f"+{int(shock*100)}%", marker_color=COLORS["profit"],
        hovertemplate="%{y}<br>ROI change: %{x:.2f}x<extra></extra>",
    ))
    fig.add_vline(x=0, line_width=1, line_color=COLORS["muted"])
    fig.update_layout(
        height=420, margin=dict(l=10, r=10, t=40, b=10),
        barmode="overlay", xaxis_title="Change in ROI (x)", yaxis_title=None,
        legend_title=None, hovermode="y unified",
    )
    return fig

def plotly_monthly_roi(df_monthly, payback_month=None, show_phased=False):
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df_monthly["Month"], y=df_monthly["Cumulative Net Revenue"],
        mode="lines+markers", name="Cumulative Net Revenue",
        line=dict(color=COLORS["revenue"], width=3),
    ))
    fig.add_trace(go.Scatter(
        x=df_monthly["Month"], y=df_monthly["Cumulative Cost"],
        mode="lines+markers", name="Cumulative Cost",
        line=dict(color=COLORS["danger"], width=3, dash="dash"),
    ))
    fig.add_trace(go.Scatter(
        x=df_monthly["Month"], y=df_monthly["Cumulative Profit"],
        mode="lines+markers", name="Cumulative Profit (Full Potential)",
        line=dict(color=COLORS["profit"], width=3),
    ))

    if show_phased and "Cumulative Phased Profit" in df_monthly.columns:
        fig.add_trace(go.Scatter(
            x=df_monthly["Month"],
            y=df_monthly["Cumulative Phased Profit"],
            mode="lines+markers",
            name="Cumulative Phased Profit",
            line=dict(color=COLORS["warning"], width=3, dash="dot"),
        ))

    if payback_month is not None:
        payback_value = df_monthly.loc[df_monthly["Month"] == payback_month, "Cumulative Profit"].iloc[0]
        fig.add_vline(x=payback_month, line_width=2, line_dash="dot", line_color=COLORS["warning"])
        fig.add_annotation(x=payback_month, y=payback_value, text=f"Payback Month: {payback_month}", showarrow=True, arrowhead=2, yshift=20)

    fig.update_layout(
        height=420, margin=dict(l=10, r=10, t=40, b=10),
        xaxis_title="Month", yaxis_title="USD", legend_title=None, hovermode="x unified",
    )
    return fig

def plotly_funnel_patients(df_funnel, tab_color):
    fig = px.bar(df_funnel, x="Patients", y="Stage", orientation="h", text="Patients")
    fig.update_traces(marker_color=tab_color, texttemplate="%{text:,.0f}", textposition="outside", cliponaxis=False)
    fig.update_layout(
        height=500, margin=dict(l=10, r=40, t=40, b=10),
        xaxis_title="Patients", yaxis_title=None, showlegend=False,
    )
    fig.update_yaxes(categoryorder="array", categoryarray=list(df_funnel["Stage"])[::-1])
    return fig

def plotly_comparison_bar(comp_df, y_col, title, y_title, color_map):
    fig = px.bar(comp_df, x="Model", y=y_col, color="Model", color_discrete_map=color_map, text=y_col)
    if "ROI" in y_col:
        fig.update_traces(texttemplate="%{text:.2f}x", textposition="outside")
    elif "Discount" in y_col:
        fig.update_traces(texttemplate="%{text:.1%}", textposition="outside")
    elif "Patients" in y_col:
        fig.update_traces(texttemplate="%{text:,.0f}", textposition="outside")
    else:
        fig.update_traces(texttemplate="$%{text:,.0f}", textposition="outside")

    fig.update_layout(
        title=title, height=320, margin=dict(l=10, r=10, t=55, b=10),
        xaxis_title=None, yaxis_title=y_title, showlegend=False,
    )
    return fig

def plotly_roi_vs_treated(comp_df, color_map):
    fig = px.scatter(
        comp_df,
        x="Treated Patients",
        y="ROI (Net)",
        color="Model",
        color_discrete_map=color_map,
        hover_name="Model",
        hover_data={
            "Treated Patients": ":,.0f",
            "Net Revenue": ":,.0f",
            "Total Cost": ":,.0f",
            "Net Profit": ":,.0f",
            "ROI (Net)": ":.2f",
            "Model": False,
        },
        text="Model",
    )
    fig.update_traces(marker=dict(size=16, line=dict(width=1, color="white")), textposition="top center")
    fig.update_layout(
        height=390, margin=dict(l=10, r=10, t=50, b=10),
        xaxis_title="Treated Patients", yaxis_title="ROI (x)", legend_title=None,
    )
    return fig

def plotly_net_profit_bar(comp_df, color_map):
    fig = px.bar(comp_df, x="Model", y="Net Profit", color="Model", color_discrete_map=color_map, text="Net Profit")
    fig.update_traces(texttemplate="$%{text:,.0f}", textposition="outside")
    fig.update_layout(
        height=360, margin=dict(l=10, r=10, t=50, b=10),
        xaxis_title=None, yaxis_title="Net Profit", showlegend=False,
    )
    return fig

def build_driver_index_df(comp_df, metrics, label_map):
    rows = []
    for metric in metrics:
        avg_val = comp_df[metric].mean()
        for _, row in comp_df.iterrows():
            indexed = (row[metric] / avg_val * 100.0) if avg_val and avg_val == avg_val else 0.0
            rows.append({
                "Model": row["Model"],
                "Metric": label_map.get(metric, metric),
                "Indexed Value": indexed,
            })
    return pd.DataFrame(rows)

def plotly_driver_index(driver_df, color_map, title):
    fig = px.bar(
        driver_df, x="Metric", y="Indexed Value", color="Model",
        barmode="group", color_discrete_map=color_map, text="Indexed Value"
    )
    fig.update_traces(texttemplate="%{text:.0f}", textposition="outside")
    fig.update_layout(
        title=title, height=380, margin=dict(l=10, r=10, t=55, b=10),
        xaxis_title=None, yaxis_title="Index (100 = selected-model average)",
        legend_title=None,
    )
    fig.add_hline(y=100, line_dash="dot", line_color=COLORS["muted"])
    return fig

def plotly_roi_vs_total_cost(comp_df, color_map):
    fig = px.scatter(
        comp_df,
        x="Total Cost",
        y="ROI (Net)",
        color="Model",
        color_discrete_map=color_map,
        hover_name="Model",
        hover_data={
            "Treated Patients": ":,.0f",
            "Net Revenue": ":,.0f",
            "Total Cost": ":,.0f",
            "Net Profit": ":,.0f",
            "ROI (Net)": ":.2f",
            "Model": False,
        },
        text="Model",
    )
    fig.update_traces(marker=dict(size=16, line=dict(width=1, color="white")), textposition="top center")
    fig.update_layout(
        height=390, margin=dict(l=10, r=10, t=50, b=10),
        xaxis_title="Total Investment", yaxis_title="ROI (x)", legend_title=None,
    )
    return fig

# -----------------------------
# Excel export helpers
# -----------------------------
def build_polished_excel_report(df_funnel, fin, colors):
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    muted_font = Font(color="6B7280")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    def set_col_widths(ws, widths):
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    def style_header_row(ws, row=1):
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = "PharmaROI Intelligence — Sponsor Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.merge_cells("A1:D1")

    summary_rows = [
        ("Treated Patients", fin["treated_patients"], "0"),
        ("Gross Revenue", fin["gross_revenue"], "$#,##0"),
        ("Discount", fin["discount"], "0.0%"),
        ("Net Revenue", fin["net_revenue"], "$#,##0"),
        ("Funnel CAC Total", fin["funnel_cac_total"], "$#,##0"),
        ("Platform Costs", fin["platform_costs_total"], "$#,##0"),
        ("Net Profit", fin["net_profit"], "$#,##0"),
        ("ROI (Net)", fin["roi_net"], "0.00x"),
    ]

    ws_sum["A3"] = "Metric"; ws_sum["B3"] = "Value"; ws_sum["C3"] = "Format"; ws_sum["D3"] = "Notes"
    style_header_row(ws_sum, 3)

    start_row = 4
    for i, (label, value, fmt) in enumerate(summary_rows):
        r = start_row + i
        ws_sum[f"A{r}"] = label
        ws_sum[f"B{r}"] = float(value) if value == value else None
        ws_sum[f"C{r}"] = fmt
        ws_sum[f"D{r}"] = ""
        ws_sum[f"A{r}"].font = bold_font if label in ("Net Revenue", "ROI (Net)", "Net Profit") else Font()
        ws_sum[f"A{r}"].alignment = left
        ws_sum[f"B{r}"].alignment = left
        ws_sum[f"C{r}"].font = muted_font
        ws_sum[f"B{r}"].number_format = fmt

    ws_sum.freeze_panes = "A4"
    set_col_widths(ws_sum, {1: 26, 2: 18, 3: 12, 4: 20})

    ws_fun = wb.create_sheet("Funnel")
    headers = list(df_funnel.columns)
    for c, h in enumerate(headers, start=1):
        ws_fun.cell(row=1, column=c, value=h)
    style_header_row(ws_fun, 1)

    for r_idx, row in enumerate(df_funnel.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws_fun.cell(row=r_idx, column=c_idx, value=val)

    ws_fun.freeze_panes = "A2"
    col_map = {name: i + 1 for i, name in enumerate(headers)}

    def fmt_col(col_name, number_format):
        if col_name not in col_map:
            return
        col = col_map[col_name]
        for rr in range(2, 2 + len(df_funnel)):
            ws_fun.cell(row=rr, column=col).number_format = number_format

    fmt_col("Patients", "0")
    fmt_col("CAC ($/pt)", "$#,##0")
    fmt_col("Stage CAC ($)", "$#,##0")
    fmt_col("Cumulative CAC ($)", "$#,##0")
    fmt_col("TAM Net Ratio", "0.00%")
    fmt_col("SAM Net Ratio", "0.00%")
    fmt_col("Net Activation Ratio", "0.00%")
    set_col_widths(ws_fun, {1: 5, 2: 52, 3: 22, 4: 12, 5: 14, 6: 12, 7: 15, 8: 18, 9: 14, 10: 14, 11: 18})

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def build_simple_excel(df, sheet_name="Data"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for i, col in enumerate(df.columns, start=1):
        max_len = max(len(str(col)), *(len(str(v)) for v in df[col].head(100).tolist())) if len(df) > 0 else len(str(col))
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 12), 28)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# -----------------------------
# Session state bootstrap
# -----------------------------
def init_session():
    if "models" not in st.session_state:
        st.session_state["models"] = [copy.deepcopy(SPONSOR_DEFAULTS)]
        st.session_state["model_names"] = ["Model 1"]
        st.session_state["active_model_idx"] = 0
    # Storage UI state
    if "storage_client_list" not in st.session_state:
        st.session_state["storage_client_list"] = None   # None = not yet fetched
    if "storage_feedback" not in st.session_state:
        st.session_state["storage_feedback"] = None
    if "confirm_delete_client" not in st.session_state:
        st.session_state["confirm_delete_client"] = False

init_session()

# -----------------------------
# Page config
# -----------------------------
st.set_page_config(page_title="PharmaROI V3 — Multi-Model", page_icon="📈", layout="wide")

# ==============================
# SIDEBAR — Client Save / Load
# ==============================
with st.sidebar:
    st.markdown("## 💾 Client Files")
    st.caption("Save or load a full set of models under a named client file.")

    # ── Refresh client list ──────────────────────────────────────────────────
    if st.button("🔄 Refresh Client List", use_container_width=True):
        st.session_state["storage_client_list"] = gs.list_clients()
        st.session_state["storage_feedback"] = None
        st.session_state["confirm_delete_client"] = False

    if st.session_state["storage_client_list"] is None:
        st.session_state["storage_client_list"] = gs.list_clients()

    client_list = st.session_state["storage_client_list"] or []

    # ── LOAD section ────────────────────────────────────────────────────────
    st.markdown("### Load a Client")
    if client_list:
        selected_client = st.selectbox(
            "Saved clients",
            options=client_list,
            key="sidebar_load_select",
            label_visibility="collapsed",
        )
        if st.button("📂 Load Client", use_container_width=True):
            payload = gs.load_client(selected_client)
            if payload:
                st.session_state["models"] = payload["models"]
                st.session_state["model_names"] = payload["model_names"]
                st.session_state["active_model_idx"] = 0
                st.session_state["storage_feedback"] = f"✅ Loaded **{selected_client}**"
                st.rerun()
    else:
        st.info("No saved clients yet.")

    st.divider()

    # ── SAVE section ────────────────────────────────────────────────────────
    st.markdown("### Save Current Work")
    save_name = st.text_input(
        "Client name",
        placeholder="e.g. Sanofi",
        key="sidebar_save_name",
        label_visibility="collapsed",
    )
    if st.button("💾 Save Client", use_container_width=True, disabled=not save_name.strip()):
        payload = {
            "models": st.session_state["models"],
            "model_names": st.session_state["model_names"],
        }
        ok = gs.save_client(save_name.strip(), payload)
        if ok:
            st.session_state["storage_client_list"] = gs.list_clients()
            st.session_state["storage_feedback"] = f"✅ Saved as **{save_name.strip()}**"
            st.rerun()

    st.divider()

    # ── DELETE section ───────────────────────────────────────────────────────
    st.markdown("### Manage Clients")
    if client_list:
        delete_target = st.selectbox(
            "Select client to delete",
            options=client_list,
            key="sidebar_delete_select",
            label_visibility="collapsed",
        )
        if not st.session_state["confirm_delete_client"]:
            if st.button("🗑️ Delete Client", use_container_width=True):
                st.session_state["confirm_delete_client"] = True
                st.rerun()
        else:
            st.warning(f"Permanently delete **{delete_target}**? This cannot be undone.")
            d_col1, d_col2 = st.columns(2)
            with d_col1:
                if st.button("Yes, Delete", use_container_width=True, type="primary"):
                    ok = gs.delete_client(delete_target)
                    if ok:
                        st.session_state["storage_client_list"] = gs.list_clients()
                        st.session_state["storage_feedback"] = f"🗑️ Deleted **{delete_target}**"
                    st.session_state["confirm_delete_client"] = False
                    st.rerun()
            with d_col2:
                if st.button("Cancel", use_container_width=True):
                    st.session_state["confirm_delete_client"] = False
                    st.rerun()
    else:
        st.caption("No clients to manage.")

    st.divider()

    # ── LOCAL FILE section ───────────────────────────────────────────────────
    with st.expander("📁 Local File Import / Export", expanded=False):
        st.caption("Download your current session as a JSON file, or upload a previously saved one.")

        # ── Download JSON ────────────────────────────────────────────────────
        st.markdown("**Download current session**")
        local_download_name = st.text_input(
            "File name (no extension)",
            placeholder="e.g. Sanofi_backup",
            key="local_download_name",
            label_visibility="collapsed",
        )
        if local_download_name.strip():
            local_payload = {
                "models": st.session_state["models"],
                "model_names": st.session_state["model_names"],
            }
            import json as _json
            json_bytes = _json.dumps(local_payload, indent=2).encode("utf-8")
            st.download_button(
                "⬇️ Download as JSON",
                data=json_bytes,
                file_name=f"{local_download_name.strip()}.json",
                mime="application/json",
                use_container_width=True,
                key="local_json_download",
            )
        else:
            st.button(
                "⬇️ Download as JSON",
                disabled=True,
                use_container_width=True,
                key="local_json_download_disabled",
                help="Enter a file name above to enable download",
            )

        st.divider()

        # ── Upload JSON ──────────────────────────────────────────────────────
        st.markdown("**Upload a saved JSON file**")
        st.caption("This will replace all current models with the contents of the file.")
        uploaded_file = st.file_uploader(
            "Choose a JSON file",
            type=["json"],
            key="local_json_upload",
            label_visibility="collapsed",
        )
        if uploaded_file is not None:
            if st.button("📂 Load from File", use_container_width=True):
                try:
                    import json as _json
                    payload = _json.loads(uploaded_file.read().decode("utf-8"))
                    if "models" in payload and "model_names" in payload:
                        st.session_state["models"] = payload["models"]
                        st.session_state["model_names"] = payload["model_names"]
                        st.session_state["active_model_idx"] = 0
                        st.session_state["storage_feedback"] = f"✅ Loaded from **{uploaded_file.name}**"
                        st.rerun()
                    else:
                        st.error("Invalid file format — make sure this is a PharmaROI JSON export.")
                except Exception as e:
                    st.error(f"Could not read file: {e}")

    # ── Feedback banner ──────────────────────────────────────────────────────
    if st.session_state["storage_feedback"]:
        st.success(st.session_state["storage_feedback"])

# -----------------------------
# Page title
# -----------------------------
st.title("PharmaROI Intelligence — V3 (Multi-Model Comparison)")
st.caption("Build multiple ROI models side-by-side and compare them in the Comparison tab.")

# -----------------------------
# Model management bar
# -----------------------------
mgmt_col1, mgmt_col2, mgmt_col3, mgmt_col4 = st.columns([2, 2, 2, 4])

with mgmt_col1:
    if st.button("➕ Add New Model", use_container_width=True):
        n = len(st.session_state["models"]) + 1
        st.session_state["models"].append(copy.deepcopy(SPONSOR_DEFAULTS))
        st.session_state["model_names"].append(f"Model {n}")
        st.session_state["active_model_idx"] = len(st.session_state["models"]) - 1
        st.rerun()

with mgmt_col2:
    copy_options = st.session_state["model_names"]
    copy_source = st.selectbox(
        "Copy from:",
        options=range(len(copy_options)),
        format_func=lambda i: copy_options[i],
        index=st.session_state["active_model_idx"],
        key="copy_source_select",
        label_visibility="collapsed",
    )
    if st.button("📋 Copy This Model", use_container_width=True):
        source_idx = copy_source
        new_state = copy.deepcopy(st.session_state["models"][source_idx])
        new_name = st.session_state["model_names"][source_idx] + " (copy)"
        st.session_state["models"].append(new_state)
        st.session_state["model_names"].append(new_name)
        st.session_state["active_model_idx"] = len(st.session_state["models"]) - 1
        st.rerun()

with mgmt_col3:
    can_delete = len(st.session_state["models"]) > 1

    if "confirm_delete" not in st.session_state:
        st.session_state["confirm_delete"] = False

    if not st.session_state["confirm_delete"]:
        if st.button("Delete Current", use_container_width=True, disabled=not can_delete):
            st.session_state["confirm_delete"] = True
            st.rerun()
    else:
        idx = st.session_state["active_model_idx"]
        st.warning(f"Delete '{st.session_state['model_names'][idx]}'?")
        confirm_cols = st.columns(2)
        with confirm_cols[0]:
            if st.button("Yes, Delete", use_container_width=True, type="primary"):
                st.session_state["models"].pop(idx)
                st.session_state["model_names"].pop(idx)
                st.session_state["active_model_idx"] = max(0, idx - 1)
                st.session_state["confirm_delete"] = False
                st.rerun()
        with confirm_cols[1]:
            if st.button("Cancel", use_container_width=True):
                st.session_state["confirm_delete"] = False
                st.rerun()

with mgmt_col4:
    idx = st.session_state["active_model_idx"]
    new_name = st.text_input(
        "Rename current model:",
        value=st.session_state["model_names"][idx],
        key=f"rename_model_{idx}",
        label_visibility="collapsed",
        placeholder="Rename current model…",
    )
    if new_name != st.session_state["model_names"][idx]:
        st.session_state["model_names"][idx] = new_name

# -----------------------------
# Tabs
# -----------------------------
tab_labels = st.session_state["model_names"] + ["Comparison"]
tabs = st.tabs(tab_labels)

for model_idx, model_tab in enumerate(tabs[:-1]):
    with model_tab:
        state = st.session_state["models"][model_idx]
        model_name = st.session_state["model_names"][model_idx]
        tab_color = TAB_PALETTE[model_idx % len(TAB_PALETTE)]

        with st.expander("Model Settings", expanded=(model_idx == st.session_state["active_model_idx"])):
            st.session_state["active_model_idx"] = model_idx

            col_r1, col_r2 = st.columns(2)
            with col_r1:
                if st.button("Reset: Sponsor Example", key=f"reset_sponsor_{model_idx}"):
                    st.session_state["models"][model_idx] = copy.deepcopy(SPONSOR_DEFAULTS)
                    st.rerun()
            with col_r2:
                if st.button("Reset: Zero", key=f"reset_zero_{model_idx}"):
                    st.session_state["models"][model_idx] = copy.deepcopy(ZERO_SAMPLE)
                    st.rerun()

            st.markdown("**Base Population**")
            state["base_population"] = st.number_input(
                "Stage 1 — Total Addressable Market",
                min_value=0, step=100_000,
                value=int(state["base_population"]),
                key=f"base_pop_{model_idx}",
            )

            st.markdown("**Revenue & Costs**")
            c1, c3 = st.columns(2)
            with c1:
                state["arpp"] = st.number_input(
                    "ARPP ($/year)",
                    min_value=0.0, step=1_000.0,
                    value=float(state["arpp"]),
                    key=f"arpp_{model_idx}",
                )
            with c3:
                state["discount"] = st.slider(
                    "Discount (gross→net)",
                    min_value=0.0, max_value=1.0, step=0.01,
                    value=float(state["discount"]),
                    key=f"discount_{model_idx}",
                )

            st.markdown("**Funnel Stages**")
            stage_names = state.get("stage_names", STAGE_NAMES[:])

            with st.expander("Customize Stage Names"):
                for sidx in range(len(STAGE_NAMES)):
                    stage_names[sidx] = st.text_input(
                        f"Stage {sidx+1} name:",
                        value=stage_names[sidx],
                        key=f"sname_{model_idx}_{sidx}",
                    )
                state["stage_names"] = stage_names

            for sidx, sname in enumerate(stage_names):
                with st.expander(f"{sidx+1}. {sname}", expanded=False):
                    state["stage_active"][sidx] = st.checkbox(
                        "Use this stage",
                        value=bool(state["stage_active"][sidx]),
                        key=f"active_{model_idx}_{sidx}",
                    )
                    if sidx == 0:
                        st.info("Stage 1 is the base population. No ratio applied.")
                    else:
                        disabled = not state["stage_active"][sidx]
                        state["ratios"][sidx] = st.slider(
                            "Funnel ratio",
                            min_value=0.0, max_value=1.0, step=0.01,
                            value=float(state["ratios"][sidx]),
                            disabled=disabled,
                            key=f"ratio_{model_idx}_{sidx}",
                        )
                    if sidx <= 5:
                        disabled = not state["stage_active"][sidx]
                        state["cac"][sidx] = st.number_input(
                            "CAC ($ per patient)",
                            min_value=0.0, step=1.0,
                            value=float(state["cac"][sidx]),
                            disabled=disabled,
                            key=f"cac_{model_idx}_{sidx}",
                        )
                    else:
                        st.caption("CAC auto-calculated from Stage 6")

            st.markdown("**Platform Costs**")
            if "platform_costs" not in state:
                state["platform_costs"] = SPONSOR_DEFAULTS["platform_costs"].copy()
            pc = state["platform_costs"]
            pc_col1, pc_col2 = st.columns(2)
            with pc_col1:
                pc["dario_connect_config"] = st.number_input("Dario Connect Configuration", min_value=0.0, step=10_000.0, value=float(pc["dario_connect_config"]), key=f"dcc_{model_idx}")
                pc["dario_care_config"] = st.number_input("Dario Care Configuration", min_value=0.0, step=10_000.0, value=float(pc["dario_care_config"]), key=f"dcarec_{model_idx}")
                pc["sub_dario_connect"] = st.number_input("Subscription — Dario Connect", min_value=0.0, step=10_000.0, value=float(pc["sub_dario_connect"]), key=f"sdc_{model_idx}")
            with pc_col2:
                pc["sub_dario_care"] = st.number_input("Subscription — Dario Care", min_value=0.0, step=10_000.0, value=float(pc["sub_dario_care"]), key=f"sdcare_{model_idx}")
                pc["maintenance_support"] = st.number_input("Maintenance & Support", min_value=0.0, step=10_000.0, value=float(pc["maintenance_support"]), key=f"ms_{model_idx}")
            st.caption(f"Total Platform Costs: {money(sum(pc.values()))}")

            st.markdown("**Optimization ROI Modeling**")
            phased_enabled = st.checkbox(
                "Enable Optimization ROI Modeling",
                value=state.get("phased_enabled", False),
                key=f"phased_enabled_{model_idx}",
            )
            state["phased_enabled"] = phased_enabled

            if phased_enabled:
                st.caption("Set revenue efficiency per optimization phase (100% = full potential).")
                ph_col1, ph_col2, ph_col3 = st.columns(3)
                with ph_col1:
                    eff_0_3_pct = st.slider(
                        "Months 0-3 efficiency",
                        min_value=0, max_value=100, step=5,
                        value=int(state.get("phased_eff_0_3", 0.33) * 100),
                        format="%d%%",
                        key=f"eff_0_3_{model_idx}",
                    )
                    eff_0_3 = eff_0_3_pct / 100
                with ph_col2:
                    eff_3_6_pct = st.slider(
                        "Months 3-6 efficiency",
                        min_value=0, max_value=100, step=5,
                        value=int(state.get("phased_eff_3_6", 0.66) * 100),
                        format="%d%%",
                        key=f"eff_3_6_{model_idx}",
                    )
                    eff_3_6 = eff_3_6_pct / 100
                with ph_col3:
                    eff_6_plus_pct = st.slider(
                        "Months 6+ efficiency",
                        min_value=0, max_value=100, step=5,
                        value=int(state.get("phased_eff_6_plus", 1.0) * 100),
                        format="%d%%",
                        key=f"eff_6_plus_{model_idx}",
                    )
                    eff_6_plus = eff_6_plus_pct / 100
                state["phased_eff_0_3"] = eff_0_3
                state["phased_eff_3_6"] = eff_3_6
                state["phased_eff_6_plus"] = eff_6_plus
            else:
                eff_0_3 = state.get("phased_eff_0_3", 0.33)
                eff_3_6 = state.get("phased_eff_3_6", 0.66)
                eff_6_plus = state.get("phased_eff_6_plus", 1.0)

        funnel_results, fin = run_model(state)
        sensitivity_df = build_roi_sensitivity_df(state, shock=0.10)

        phased_enabled = state.get("phased_enabled", False)
        _eff_0_3 = state.get("phased_eff_0_3", 0.2) if phased_enabled else 1.0
        _eff_3_6 = state.get("phased_eff_3_6", 0.6) if phased_enabled else 1.0
        _eff_6_plus = state.get("phased_eff_6_plus", 1.0) if phased_enabled else 1.0

        if pd is not None:
            df_monthly, payback_month, phased_net_revenue, phased_roi = build_monthly_roi_df(
                fin, state,
                eff_0_3=_eff_0_3,
                eff_3_6=_eff_3_6,
                eff_6_plus=_eff_6_plus,
            )
        else:
            df_monthly, payback_month, phased_net_revenue, phased_roi = None, None, 0.0, float("nan")

        tam_patients = funnel_results[0].patients
        sam_patients = funnel_results[1].patients
        activation_patients = funnel_results[5].patients

        st.markdown(
            f"<div style='border-left: 4px solid {tab_color}; padding-left: 12px; margin-bottom: 8px;'><strong style='font-size:1.1rem'>{model_name}</strong></div>",
            unsafe_allow_html=True,
        )

        roi = fin["roi_net"]
        total_cost = fin["funnel_cac_total"] + fin["platform_costs_total"]

        k1, k2, k3, k4, k5, k6 = st.columns(6)
        k1.metric("ROI (Net)", roix(roi) if roi == roi else "—")
        k2.metric("Treated Patients", number(fin["treated_patients"]))
        k3.metric("Net Revenue", money(fin["net_revenue"]))
        k4.metric("Funnel CAC", money(fin["funnel_cac_total"]))
        k5.metric("Total Cost", money(total_cost))
        k6.metric("Net Profit", money(fin["net_profit"]))

        st.markdown(
            f"Gross: **\\${fin['gross_revenue']:,.0f}**  |  "
            f"Discount: **{fin['discount']*100:.1f}%**  |  "
            f"Discount Amount: **\\${fin['gross_revenue'] - fin['net_revenue']:,.0f}**  |  "
            f"Net Revenue per Rx: **\\${(float(state['arpp']) * (1 - fin['discount'])):,.0f}**"
        )

        if phased_enabled:
            st.markdown("### Phased Optimization Outlook")
            total_cost = fin["funnel_cac_total"] + fin["platform_costs_total"]
            monthly_net_revenue = fin["net_revenue"] / max(1, int(round(float(state["treatment_years"]) * 12)))

            roi_0_3 = roi * _eff_0_3
            roi_3_6 = roi * _eff_3_6
            roi_6_plus = roi * _eff_6_plus

            ph1, ph2, ph3 = st.columns(3)
            ph1.metric(
                "ROI — Months 0-3",
                roix(roi_0_3) if roi_0_3 == roi_0_3 else "—",
                delta=f"{_eff_0_3:.0%} efficiency",
                delta_color="off",
            )
            ph2.metric(
                "ROI — Months 3-6",
                roix(roi_3_6) if roi_3_6 == roi_3_6 else "—",
                delta=f"{_eff_3_6:.0%} efficiency",
                delta_color="off",
            )
            ph3.metric(
                "ROI — Months 6+",
                roix(roi_6_plus) if roi_6_plus == roi_6_plus else "—",
                delta=f"{_eff_6_plus:.0%} efficiency",
                delta_color="off",
            )
            st.caption(
                f"ROI shown per phase window against total cost. Full Potential ROI: **{roix(roi)}**"
            )

        st.subheader("Funnel Table")
        table_rows = []
        for ridx, r in enumerate(funnel_results):
            tam_ratio = r.patients / tam_patients if tam_patients > 0 else 0.0
            sam_ratio = r.patients / sam_patients if sam_patients > 0 else 0.0
            net_activation = r.patients / activation_patients if activation_patients > 0 else 0.0

            table_rows.append({
                "#": ridx + 1,
                "Stage": r.name,
                "Status": "Active" if r.active else "Inactive (pass-through)",
                "Ratio Used": "—" if ridx == 0 else pct(r.ratio_used),
                "Patients": float(r.patients),
                "CAC ($/pt)": float(r.cac_per_patient),
                "Stage CAC ($)": float(r.stage_cac),
                "Cumulative CAC ($)": float(r.cumulative_cac),
                "TAM Net Ratio": float(tam_ratio),
                "SAM Net Ratio": float(sam_ratio),
                "Net Activation Ratio": float(net_activation),
            })

        if pd is not None:
            df_funnel = pd.DataFrame(table_rows)
            df_display = df_funnel.copy()
            df_display["Patients"] = df_display["Patients"].map(lambda x: f"{x:,.0f}")
            df_display["CAC ($/pt)"] = df_display["CAC ($/pt)"].map(lambda x: f"${x:,.0f}")
            df_display["Stage CAC ($)"] = df_display["Stage CAC ($)"].map(lambda x: f"${x:,.0f}")
            df_display["Cumulative CAC ($)"] = df_display["Cumulative CAC ($)"].map(lambda x: f"${x:,.0f}")
            df_display["TAM Net Ratio"] = df_display["TAM Net Ratio"].map(lambda x: "—" if x > 1.0 else f"{x*100:.2f}%")
            df_display["SAM Net Ratio"] = df_display["SAM Net Ratio"].map(lambda x: "—" if x > 1.0 else f"{x*100:.2f}%")
            df_display["Net Activation Ratio"] = df_display["Net Activation Ratio"].map(lambda x: "—" if x > 1.0 else f"{x*100:.2f}%")
            st.dataframe(df_display, use_container_width=True, hide_index=True)

            st.markdown("### Export")
            ec1, ec2 = st.columns(2)
            with ec1:
                xlsx_bytes = build_polished_excel_report(df_funnel, fin, COLORS)
                st.download_button(
                    "⬇️ Download Excel Report",
                    data=xlsx_bytes,
                    file_name=f"{model_name.replace(' ', '_')}_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_xlsx_{model_idx}",
                )
            with ec2:
                csv_data = df_funnel.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "⬇️ Download CSV",
                    data=csv_data,
                    file_name=f"{model_name.replace(' ', '_')}_funnel.csv",
                    mime="text/csv",
                    key=f"dl_csv_{model_idx}",
                )
        else:
            st.write(table_rows)

        st.subheader("Visuals")
        chart_row1_col1, chart_row1_col2 = st.columns(2)

        with chart_row1_col1:
            st.markdown("**Revenue / Cost / Profit Waterfall**")
            st.plotly_chart(plotly_waterfall(fin), use_container_width=True)

        with chart_row1_col2:
            st.markdown("**ROI Sensitivity**")
            if pd is not None and sensitivity_df is not None:
                st.plotly_chart(plotly_sensitivity_tornado(sensitivity_df, shock=0.10), use_container_width=True)
            else:
                st.info("Sensitivity chart requires pandas.")

        if pd is not None and df_monthly is not None:
            with st.expander("Monthly ROI Comparison", expanded=True):
                st.caption(
                    "Net revenue spread evenly across treatment months. "
                    + ("Phased Cumulative Profit shown as a dotted line." if phased_enabled else "Enable Phased ROI Modeling to overlay a ramp-up curve.")
                )
                st.plotly_chart(
                    plotly_monthly_roi(df_monthly, payback_month, show_phased=phased_enabled),
                    use_container_width=True,
                )

        with st.expander("Funnel Visualization", expanded=True):
            if pd is not None:
                funnel_chart_df = pd.DataFrame([{"Stage": r.name, "Patients": r.patients} for r in funnel_results])
                st.plotly_chart(plotly_funnel_patients(funnel_chart_df, tab_color), use_container_width=True)

# -----------------------------
# Comparison Tab
# -----------------------------
with tabs[-1]:
    st.subheader("Model Comparison")

    if len(st.session_state["models"]) < 2:
        st.info("Add at least 2 models to compare them here.")
    else:
        st.markdown("**Select models to compare:**")
        selected_model_names = st.multiselect(
            "Choose models",
            options=st.session_state["model_names"],
            default=st.session_state["model_names"],
            key="comparison_model_select",
            label_visibility="collapsed",
        )

        if len(selected_model_names) < 2:
            st.warning("Please select at least 2 models to compare.")
            st.stop()

        selected_indices = [i for i, name in enumerate(st.session_state["model_names"]) if name in selected_model_names]
        selected_models = [st.session_state["models"][i] for i in selected_indices]
        selected_names = [st.session_state["model_names"][i] for i in selected_indices]

        comparison_rows = []
        monthly_rows = []

        for mstate, mname in zip(selected_models, selected_names):
            _, fin = run_model(mstate)
            roi = fin["roi_net"]
            total_cost = fin["funnel_cac_total"] + fin["platform_costs_total"]

            comparison_rows.append({
                "Model": mname,
                "Treated Patients": fin["treated_patients"],
                "Gross Revenue": fin["gross_revenue"],
                "Net Revenue": fin["net_revenue"],
                "Funnel CAC": fin["funnel_cac_total"],
                "Platform Costs": fin["platform_costs_total"],
                "Total Cost": total_cost,
                "Net Profit": fin["net_profit"],
                "Discount": fin["discount"],
                "ARPP": float(mstate["arpp"]),
                "ROI (Net)": roi if roi == roi else 0.0,
            })

            if pd is not None:
                monthly_df, _, _phased_rev, _phased_roi = build_monthly_roi_df(fin, mstate)
                monthly_copy = monthly_df.copy()
                monthly_copy["Model"] = mname
                monthly_rows.append(monthly_copy)

        if pd is not None:
            comp_df = pd.DataFrame(comparison_rows)

            color_map = {name: TAB_PALETTE[i % len(TAB_PALETTE)] for i, name in enumerate(selected_names)}

            st.markdown("### Key Metrics")
            disp = comp_df.copy()
            disp["Treated Patients"] = disp["Treated Patients"].map(lambda x: f"{x:,.0f}")
            disp["Gross Revenue"] = disp["Gross Revenue"].map(lambda x: f"${x:,.0f}")
            disp["Net Revenue"] = disp["Net Revenue"].map(lambda x: f"${x:,.0f}")
            disp["Funnel CAC"] = disp["Funnel CAC"].map(lambda x: f"${x:,.0f}")
            disp["Platform Costs"] = disp["Platform Costs"].map(lambda x: f"${x:,.0f}")
            disp["Total Cost"] = disp["Total Cost"].map(lambda x: f"${x:,.0f}")
            disp["Net Profit"] = disp["Net Profit"].map(lambda x: f"${x:,.0f}")
            disp["Discount"] = disp["Discount"].map(lambda x: f"{x*100:.1f}%")
            disp["ARPP"] = disp["ARPP"].map(lambda x: f"${x:,.0f}")
            disp["ROI (Net)"] = disp["ROI (Net)"].map(lambda x: f"{x:.2f}x")
            st.dataframe(disp, use_container_width=True, hide_index=True)

            st.markdown("### Charts")
            chart_col1, chart_col2 = st.columns(2)

            with chart_col1:
                st.plotly_chart(
                    plotly_comparison_bar(comp_df, "ROI (Net)", "ROI (Net)", "ROI (x)", color_map),
                    use_container_width=True,
                )

            with chart_col2:
                st.plotly_chart(
                    plotly_comparison_bar(comp_df, "Net Revenue", "Net Revenue", "USD", color_map),
                    use_container_width=True,
                )

            chart_col3, chart_col4 = st.columns(2)

            with chart_col3:
                st.plotly_chart(
                    plotly_comparison_bar(comp_df, "Treated Patients", "Treated Patients", "Patients", color_map),
                    use_container_width=True,
                )

            with chart_col4:
                st.plotly_chart(
                    plotly_comparison_bar(comp_df, "Total Cost", "Total Cost", "USD", color_map),
                    use_container_width=True,
                )

            if monthly_rows:
                st.markdown("### Monthly ROI Comparison")
                monthly_comp_df = pd.concat(monthly_rows, ignore_index=True)

                fig_monthly_comp = px.line(
                    monthly_comp_df,
                    x="Month",
                    y="Cumulative Profit",
                    color="Model",
                    markers=True,
                    color_discrete_map=color_map,
                )
                fig_monthly_comp.update_layout(
                    height=380,
                    margin=dict(l=10, r=10, t=40, b=10),
                    xaxis_title="Month",
                    yaxis_title="Cumulative Profit",
                    hovermode="x unified",
                    legend_title=None,
                )
                st.plotly_chart(fig_monthly_comp, use_container_width=True)

            st.markdown("### ROI Insights")

            st.markdown("#### ROI vs Patient Impact")
            st.caption("Compares scenario efficiency and patient impact on the same view.")
            roi_vs_patient_df = comp_df[["Model", "Treated Patients", "Net Revenue", "Total Cost", "Net Profit", "ROI (Net)"]].copy()
            st.plotly_chart(plotly_roi_vs_treated(comp_df, color_map), use_container_width=True)
            st.download_button(
                "Download Data (Excel)",
                data=build_simple_excel(roi_vs_patient_df, "ROI vs Patient Impact"),
                file_name="roi_vs_patient_impact.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_roi_vs_patient_impact",
            )

            st.markdown("#### Net Profit by Scenario")
            st.caption("Shows absolute financial value after funnel and platform costs are applied.")
            net_profit_df = comp_df[["Model", "Net Profit", "Net Revenue", "Total Cost", "ROI (Net)"]].copy()
            st.plotly_chart(plotly_net_profit_bar(comp_df, color_map), use_container_width=True)
            st.download_button(
                "Download Data (Excel)",
                data=build_simple_excel(net_profit_df, "Net Profit by Scenario"),
                file_name="net_profit_by_scenario.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_net_profit_by_scenario",
            )

            st.markdown("#### Key Financial Drivers Comparison")
            st.caption("Compares major revenue and cost drivers relative to the average across the selected models.")
            revenue_metrics = ["ARPP", "Discount", "Treated Patients"]
            revenue_label_map = {"ARPP": "ARPP", "Discount": "Discount", "Treated Patients": "Treated Patients"}
            cost_metrics = ["Funnel CAC", "Platform Costs", "Total Cost"]
            cost_label_map = {"Funnel CAC": "Funnel CAC", "Platform Costs": "Platform Costs", "Total Cost": "Total Cost"}

            revenue_driver_df = build_driver_index_df(comp_df, revenue_metrics, revenue_label_map)
            cost_driver_df = build_driver_index_df(comp_df, cost_metrics, cost_label_map)
            driver_export_df = pd.concat([
                revenue_driver_df.assign(Section="Revenue Drivers"),
                cost_driver_df.assign(Section="Cost Drivers"),
            ], ignore_index=True)[["Section", "Model", "Metric", "Indexed Value"]]

            st.download_button(
                "Download Data (Excel)",
                data=build_simple_excel(driver_export_df, "Financial Drivers"),
                file_name="financial_driver_comparison.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_financial_driver_comparison",
            )

            d1, d2 = st.columns(2)

            with d1:
                st.markdown("##### Revenue Drivers")
                st.plotly_chart(
                    plotly_driver_index(revenue_driver_df, color_map, "Revenue Drivers"),
                    use_container_width=True,
                )
                st.download_button(
                    "Download Data (Excel)",
                    data=build_simple_excel(revenue_driver_df, "Revenue Drivers"),
                    file_name="revenue_drivers.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_revenue_drivers",
                )

            with d2:
                st.markdown("##### Cost Drivers")
                st.plotly_chart(
                    plotly_driver_index(cost_driver_df, color_map, "Cost Drivers"),
                    use_container_width=True,
                )
                st.download_button(
                    "Download Data (Excel)",
                    data=build_simple_excel(cost_driver_df, "Cost Drivers"),
                    file_name="cost_drivers.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_cost_drivers",
                )

            st.markdown("#### ROI vs Total Investment")
            st.caption("Shows how efficiently each scenario converts total investment into ROI.")
            roi_vs_investment_df = comp_df[["Model", "Total Cost", "Treated Patients", "Net Revenue", "Net Profit", "ROI (Net)"]].copy()
            st.plotly_chart(plotly_roi_vs_total_cost(comp_df, color_map), use_container_width=True)
            st.download_button(
                "Download Data (Excel)",
                data=build_simple_excel(roi_vs_investment_df, "ROI vs Total Investment"),
                file_name="roi_vs_total_investment.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_roi_vs_total_investment",
            )

            st.markdown("### Model Diff View")
            if len(selected_names) >= 2:
                diff_col1, diff_col2 = st.columns(2)
                with diff_col1:
                    diff_model_a = st.selectbox("Model A:", options=selected_names, index=0, key="diff_model_a")
                with diff_col2:
                    remaining = [n for n in selected_names if n != diff_model_a]
                    diff_model_b = st.selectbox("Model B:", options=remaining, index=0, key="diff_model_b")

                idx_a = st.session_state["model_names"].index(diff_model_a)
                idx_b = st.session_state["model_names"].index(diff_model_b)
                state_a = st.session_state["models"][idx_a]
                state_b = st.session_state["models"][idx_b]

                diff_rows = []

                top_params = [
                    ("Base Population", "base_population", "{:,.0f}"),
                    ("ARPP", "arpp", "${:,.0f}"),
                    ("Treatment Years", "treatment_years", "{:.1f}"),
                    ("Discount", "discount", "{:.1%}"),
                ]
                for label, key, fmt in top_params:
                    val_a = state_a.get(key, 0)
                    val_b = state_b.get(key, 0)
                    if val_a != val_b:
                        diff_rows.append({
                            "Parameter": label,
                            f"{diff_model_a}": fmt.format(val_a),
                            f"{diff_model_b}": fmt.format(val_b),
                            "Difference": fmt.format(val_b - val_a) if "%" not in fmt else f"{(val_b - val_a)*100:+.1f}pp",
                        })

                for sidx in range(len(STAGE_NAMES)):
                    ratio_a = state_a["ratios"][sidx]
                    ratio_b = state_b["ratios"][sidx]
                    if ratio_a != ratio_b and sidx > 0:
                        diff_rows.append({
                            "Parameter": f"Stage {sidx+1} Ratio",
                            f"{diff_model_a}": f"{ratio_a:.1%}",
                            f"{diff_model_b}": f"{ratio_b:.1%}",
                            "Difference": f"{(ratio_b - ratio_a)*100:+.1f}pp",
                        })

                    cac_a = state_a["cac"][sidx]
                    cac_b = state_b["cac"][sidx]
                    if cac_a != cac_b:
                        diff_rows.append({
                            "Parameter": f"Stage {sidx+1} CAC",
                            f"{diff_model_a}": f"${cac_a:,.0f}",
                            f"{diff_model_b}": f"${cac_b:,.0f}",
                            "Difference": f"${cac_b - cac_a:+,.0f}",
                        })

                if diff_rows:
                    diff_df = pd.DataFrame(diff_rows)
                    st.dataframe(diff_df, use_container_width=True, hide_index=True)
                else:
                    st.success("These two models have identical parameters!")
            else:
                st.info("Select at least 2 models above to see a diff view.")

            st.markdown("### Export Comparison")
            comp_csv = comp_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Download Comparison CSV",
                data=comp_csv,
                file_name="pharmaroi_comparison.csv",
                mime="text/csv",
            )

st.divider()
st.subheader("How to interpret")
st.write("""
- Each **model tab** is fully independent — tweak funnel stages, ratios, CAC, ARPP, and discount separately.
- Use **Add New Model** or **Duplicate Current** to create variants.
- The **Comparison** tab shows all models side-by-side with charts and a downloadable table.
- Use the **sidebar** to save your current work under a client name, or load a previously saved client.
- **ROI (Net)** = Net Revenue / (Funnel CAC + Platform Costs)
- **Net Profit** = Net Revenue − Funnel CAC − Platform Costs
- **Net Revenue** = Gross Revenue × (1 − Discount)
- **TAM Net Ratio** = Patients at Stage / Stage 1 (Total Addressable Market)
- **SAM Net Ratio** = Patients at Stage / Stage 2 (F2 and F3)
- **Net Activation Ratio** = Patients at Stage / Stage 6 (Activation onto Dario Connect)

**ROI Sensitivity assumption:**
- Sensitivity uses one-at-a-time shocks of ±10% on selected variables.

**Monthly ROI Comparison assumption:**
- Net revenue is spread evenly across treatment duration months.
- Funnel CAC and platform costs are treated as upfront Month 1 costs for that chart only.
""")
